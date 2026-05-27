"""
tinvest_to_finance.py
---------------------
Two modes, auto-detected:

  CREATE  — Finance_updated.xlsx does not exist → build from scratch.
  UPDATE  — file already exists → patch only price/qty/cashflow columns;
            all risk, geography and qualitative columns are untouched.

Extra commands (pass as any argument):
  charts      — pull live portfolio and render fancy distribution charts
  correlation — pull historical prices and render correlation matrix
  dcf         — print a DCF analysis template skeleton

Usage:
    export INVEST_TOKEN='your_token'

    python tinvest_to_finance.py                              # update/create xlsx + charts
    python tinvest_to_finance.py charts                       # only charts
    python tinvest_to_finance.py correlation                  # correlation matrix (180d)
    python tinvest_to_finance.py correlation 365              # correlation matrix (365d)
    python tinvest_to_finance.py correlation 180 spearman     # Spearman instead of Pearson
    python tinvest_to_finance.py correlation top15            # top 15 holdings by weight
    python tinvest_to_finance.py correlation min1             # exclude positions below 1% weight
    python tinvest_to_finance.py dcf                          # DCF template (no API needed)
    python tinvest_to_finance.py MyPortfolio.xlsx             # custom file path
"""

import os
import sys
import logging
from decimal import Decimal
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
import matplotlib.ticker as mticker
from openpyxl import Workbook

from correlation_matrix import run_correlation_analysis
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from t_tech.invest import Client

logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

TOKEN = os.getenv("INVEST_TOKEN")
if not TOKEN:
    raise ValueError(
        "Не задана переменная окружения INVEST_TOKEN.\n"
        "Установите её командой: export INVEST_TOKEN='ваш_токен'"
    )

# ═══════════════════════════════════════════════════════════════════════════════
# API HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

CURRENCY_FIGI = {
    'usd': 'BBG0013HGFT4',
    'eur': 'BBG0013HJJ31',
    'cny': 'BBG0013HRTL0',
    'hkd': 'BBG0013HSW87',
    'chf': 'BBG0013HQ5K4',
}


def _q(q) -> Decimal:
    """Convert Quotation proto to Decimal."""
    return Decimal(q.units) + Decimal(q.nano) / Decimal(1_000_000_000)


def get_currency_rates(client: Client) -> Dict[str, Decimal]:
    rates: Dict[str, Decimal] = {'rub': Decimal('1')}
    for curr, figi in CURRENCY_FIGI.items():
        try:
            resp = client.market_data.get_last_prices(figi=[figi])
            if resp.last_prices:
                rates[curr] = _q(resp.last_prices[0].price)
        except Exception as e:
            logger.warning(f"Не удалось получить курс {curr.upper()}: {e}")
            rates[curr] = Decimal('1')
    return rates


def get_instrument_info(client: Client, figi: str) -> Optional[Dict]:
    try:
        resp = client.instruments.find_instrument(query=figi)
        for inst in resp.instruments:
            inst_figi = getattr(inst, 'figi', getattr(inst, 'id', None))
            if inst_figi == figi:
                name  = getattr(inst, 'name', 'Неизвестно')
                itype = getattr(inst, 'instrument_type', 'N/A')
                if any(x in name.lower() for x in ('золот', 'gold', 'metal')):
                    itype = 'precious_metal'
                return {
                    'name':            name,
                    'ticker':          getattr(inst, 'ticker', figi[:10]),
                    'currency':        getattr(inst, 'currency', 'N/A'),
                    'instrument_type': itype,
                }
    except Exception as e:
        logger.error(f"Ошибка при получении инструмента {figi}: {e}")
    return None


def _is_option(row: dict) -> bool:
    """Options: unknown name AND negative quantity — skip them."""
    name    = row.get('name', '')
    qty     = row.get('quantity', Decimal('0'))
    unknown = name in ('', 'Неизвестный инструмент', 'Неизвестно', 'N/A')
    return unknown and qty < 0


def fetch_all_positions(client: Client,
                        rates: Dict[str, Decimal]) -> List[Dict]:
    """Return flat list of position dicts across all accounts, excluding options."""
    accounts = client.users.get_accounts().accounts
    rows: List[Dict] = []

    for account in accounts:
        try:
            portfolio = client.operations.get_portfolio(account_id=account.id)
        except Exception as e:
            logger.error(f"Ошибка портфеля для счёта {account.id}: {e}")
            continue

        for pos in portfolio.positions:
            info = get_instrument_info(client, pos.figi) or {
                'name':            'Неизвестный инструмент',
                'ticker':          pos.figi[:10],
                'currency':        'N/A',
                'instrument_type': 'N/A',
            }

            qty          = _q(pos.quantity) if pos.quantity else Decimal('0')
            avg_price    = _q(pos.average_position_price) if pos.average_position_price else None
            cur_price    = None
            cur_currency = 'rub'
            if pos.current_price:
                cur_price    = _q(pos.current_price)
                cur_currency = pos.current_price.currency.lower()

            cashflow_year = _q(pos.expected_yield) if pos.expected_yield else None

            # RUB value for charting
            rub_value = Decimal('0')
            if cur_price and qty > 0:
                val = cur_price * qty
                rub_value = val * rates.get(cur_currency, Decimal('1')) \
                            if cur_currency != 'rub' else val

            row = {
                'account':            account.name,
                'figi':               pos.figi,
                'name':               info['name'],
                'ticker':             info['ticker'],
                'instrument_type':    info['instrument_type'],
                'quantity':           qty,
                'avg_price':          avg_price,
                'avg_price_currency': cur_currency,
                'cur_price':          cur_price,
                'cur_price_currency': cur_currency,
                'cashflow_year':      cashflow_year,
                'rub_value':          rub_value,
            }

            if _is_option(row):
                logger.info(f"  Пропущен опцион: {pos.figi}")
                continue

            rows.append(row)

    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL — shared style helpers
# ═══════════════════════════════════════════════════════════════════════════════

THIN        = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', fgColor='BDD7EE')
RISK_FILL   = PatternFill('solid', fgColor='FCE4D6')
LEGEND_FILL = PatternFill('solid', fgColor='E2EFDA')

COL_WIDTHS = {
    1: 12, 2: 28, 3: 10, 4: 18, 5: 12, 6: 14, 7: 14, 8: 16,
    9: 15, 10: 22, 11: 18, 12: 22, 13: 22, 14: 28, 15: 30,
    16: 30, 17: 10, 18: 10, 19: 12, 20: 14, 21: 11, 22: 10,
    23: 12, 24: 14, 25: 10, 26: 40, 27: 30, 28: 20, 29: 40, 30: 40,
}

HEADERS_ROW2 = {
    'B': 'Инфляция рубля: 15%', 'D': 'Капитал',
    'L': 'География актива',    'O': 'Экономическая позиция ',
    'Q': 'Риски',               'AA': 'Почему',
    'AB': 'Наличие страховки ', 'AC': 'Причина добавления в портфель',
    'AD': 'Когда продавать',
}

HEADERS_ROW3 = [
    'Счет', 'Название', 'Тикер', 'Тип инструмента ', 'Количество ',
    'Цена покупки ', 'Цена сейчас', 'Общая стоимость', 'Кэшфлоу(год)',
    'Ожидаемые дивиденты/ак', 'Ежемесячный доход',
    'Основная страна бизнеса', 'Страна хранения актива',
    'Валюта в которой работает актив', 'Макро', 'Микро',
    'Ценовые', 'Страновые', 'Регуляторные ', 'Бизнес-риски',
    'Кредитные ', 'Валютные ', 'Инфраструк.', 'Инфляция влад.', 'Товарные ',
    'Риски детально', 'Цель актива', 'Наличие страховки',
    'Причина добавления в портфель', 'Когда продавать',
]

ROW2_MERGES = [('D2', 'K2'), ('L2', 'N2'), ('O2', 'P2'), ('Q2', 'Z2')]

RISK_LEGEND = [
    (34, 'Ценовые(компания и ее акции могут падать вместе с рынком и вообще акция по любым причинам может стоить сколько угодно)'),
    (35, 'Страновые(в стране жопа и с активом жопа)'),
    (36, 'Регуляторные(запретили что-то и компании жопа)'),
    (37, 'Бизнес-риски(как и у любого риска, что-то пойдет не так в операционке)'),
    (38, 'Кредитные риски/риск дефолта(заплятят ли по долгам)'),
    (39, 'Валютные(импортер/экспортер)'),
    (40, 'Инфраструтурные(санкции и прочая хуйня для инвестора в РФ)'),
    (41, 'Товарные(риск сосредоточен на цене определенного коммодика)'),
    (43, 'Шкала'), (44, 'Нет риска'), (45, 'Чуть больше'),
    (46, 'Еще чуть-чуть больше'), (47, 'Средний риск'),
    (48, 'Чуть выше среднего'), (49, 'Еще чуть-чуть выше'), (50, 'Риск велик'),
]
RISK_LEGEND_VALUES = {44: 1, 45: 2, 46: 3, 47: 4, 48: 5, 49: 6, 50: 7}


def _hf():    return Font(name='Arial', bold=True, size=10)
def _df():    return Font(name='Arial', size=10)
def _wrap():  return Alignment(wrap_text=True, vertical='top')
def _center():return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _apply_data_row_style(ws, r: int):
    for col in range(1, 31):
        cell           = ws.cell(row=r, column=col)
        cell.font      = _df()
        cell.alignment = _wrap()
        cell.border    = THIN_BORDER
        if 17 <= col <= 25:
            cell.fill = RISK_FILL
    for col in (5, 6, 7, 8, 9, 10, 11):
        ws.cell(row=r, column=col).number_format = '#,##0.00'


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL — CREATE
# ═══════════════════════════════════════════════════════════════════════════════

def build_workbook(rows: List[Dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Assets'

    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 43
    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_letter, label in HEADERS_ROW2.items():
        cell           = ws[f'{col_letter}2']
        cell.value     = label
        cell.font      = _hf()
        cell.fill      = HEADER_FILL
        cell.alignment = _center()
    for a, b in ROW2_MERGES:
        ws.merge_cells(f'{a}:{b}')

    for col_idx, header in enumerate(HEADERS_ROW3, start=1):
        cell           = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = _hf()
        cell.fill      = HEADER_FILL
        cell.alignment = _center()
        cell.border    = THIN_BORDER

    for r_idx, pos in enumerate(rows, start=4):
        ws.row_dimensions[r_idx].height = 60
        ws.cell(r_idx, 1).value = pos['account']
        ws.cell(r_idx, 2).value = pos['name']
        ws.cell(r_idx, 3).value = pos['ticker']
        ws.cell(r_idx, 4).value = pos['instrument_type']
        ws.cell(r_idx, 5).value = float(pos['quantity'])
        if pos['avg_price'] is not None:
            ws.cell(r_idx, 6).value = float(pos['avg_price'])
        if pos['cur_price'] is not None:
            ws.cell(r_idx, 7).value = float(pos['cur_price'])
        ws.cell(r_idx, 8).value  = f'=E{r_idx}*G{r_idx}'
        if pos['cashflow_year'] is not None:
            ws.cell(r_idx, 9).value = float(pos['cashflow_year'])
        ws.cell(r_idx, 11).value = f'=J{r_idx}*E{r_idx}/12'
        _apply_data_row_style(ws, r_idx)

    last = 3 + len(rows)
    ss   = last + 2
    ws.cell(ss,     11).value = 'Ежемесячный доход:'
    ws.cell(ss,     12).value = f'=SUM(K4:K{last})'
    ws.cell(ss,     13).value = f'=SUM(K4:K{last})/SUM(H4:H{last})*100'
    ws.cell(ss + 1, 11).value = f'=SUM(K4:K{last})'
    ws.cell(ss + 1, 12).value = f'=SUM(H4:H{last})'
    ws.cell(ss + 2,  2).value = 'Риски:'
    ws.cell(ss + 2, 12).value = f'=SUM(H4:H{last})'
    for row in (ss, ss + 1, ss + 2):
        ws.row_dimensions[row].height = 15

    lo = ss + 4
    for template_row, text in RISK_LEGEND:
        ar   = lo + (template_row - 34)
        cell = ws.cell(ar, 2, value=text)
        cell.font = _df(); cell.alignment = _wrap(); cell.fill = LEGEND_FILL
        if template_row in RISK_LEGEND_VALUES:
            vc = ws.cell(ar, 4, value=RISK_LEGEND_VALUES[template_row])
            vc.font = _df(); vc.alignment = _center(); vc.fill = LEGEND_FILL

    for name in ('Speculative', 'Deposit', 'Transactions'):
        wb.create_sheet(name)
    return wb


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL — UPDATE  (de-duplicate fix)
# ═══════════════════════════════════════════════════════════════════════════════

def update_workbook(path: str, api_rows: List[Dict]) -> None:
    """
    Patch existing file in-place.
    FIX: de-duplicate api_rows by (account, ticker) before processing —
         this prevents the doubling bug where the same row was written twice.
    """
    from openpyxl import load_workbook as _load

    # De-duplicate: keep last value per (account, ticker) key
    seen: Dict[tuple, Dict] = {}
    for pos in api_rows:
        key = (str(pos['account']).strip(), str(pos['ticker']).strip())
        seen[key] = pos
    deduped = list(seen.values())
    logger.info(f"API позиций после дедупликации: {len(deduped)}")

    wb = _load(path)
    ws = wb['Assets']

    # Build lookup: (account, ticker) -> first row index
    existing: Dict[tuple, int] = {}
    last_data_row = 3
    for row in ws.iter_rows(min_row=4):
        acct_val   = row[0].value
        ticker_val = row[2].value
        if ticker_val is not None:
            key = (str(acct_val).strip() if acct_val else '',
                   str(ticker_val).strip())
            if key not in existing:          # first occurrence only
                existing[key] = row[0].row
            last_data_row = row[0].row

    logger.info(f"Существующих строк в файле: {len(existing)}")
    updated = appended = 0

    for pos in deduped:
        key = (str(pos['account']).strip(), str(pos['ticker']).strip())

        if key in existing:
            r = existing[key]
            if pos['quantity']      is not None: ws.cell(r, 5).value = float(pos['quantity'])
            if pos['avg_price']     is not None: ws.cell(r, 6).value = float(pos['avg_price'])
            if pos['cur_price']     is not None: ws.cell(r, 7).value = float(pos['cur_price'])
            ws.cell(r, 8).value  = f'=E{r}*G{r}'
            if pos['cashflow_year'] is not None: ws.cell(r, 9).value = float(pos['cashflow_year'])
            ws.cell(r, 11).value = f'=J{r}*E{r}/12'
            updated += 1
            logger.info(f"  Обновлено: {key}")
        else:
            last_data_row += 1
            r = last_data_row
            ws.row_dimensions[r].height = 60
            ws.cell(r, 1).value = pos['account']
            ws.cell(r, 2).value = pos['name']
            ws.cell(r, 3).value = pos['ticker']
            ws.cell(r, 4).value = pos['instrument_type']
            if pos['quantity']      is not None: ws.cell(r, 5).value = float(pos['quantity'])
            if pos['avg_price']     is not None: ws.cell(r, 6).value = float(pos['avg_price'])
            if pos['cur_price']     is not None: ws.cell(r, 7).value = float(pos['cur_price'])
            ws.cell(r, 8).value  = f'=E{r}*G{r}'
            if pos['cashflow_year'] is not None: ws.cell(r, 9).value = float(pos['cashflow_year'])
            ws.cell(r, 11).value = f'=J{r}*E{r}/12'
            _apply_data_row_style(ws, r)
            existing[key] = r
            appended += 1
            logger.info(f"  Добавлено: {key}")

    logger.info(f"Итого: обновлено {updated}, добавлено {appended}.")
    wb.save(path)
    logger.info(f"Файл сохранён: {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# CHARTS  — dark Bloomberg-style financial dashboard
# ═══════════════════════════════════════════════════════════════════════════════

_DARK_BG  = '#0D1117'
_PANEL_BG = '#161B22'
_ACCENT1  = '#58A6FF'
_TEXT     = '#E6EDF3'
_DIM      = '#8B949E'
_GRID     = '#21262D'

TYPE_PALETTE = {
    'Акции':         '#58A6FF',
    'Облигации':     '#3FB950',
    'Фонды':         '#D29922',
    'Драг. металлы': '#F0C040',
    'Недвижимость':  '#BC8CFF',
    'Крипта':        '#F78166',
    'Деньги':        '#6E7681',
    'Другие':        '#484F58',
}

TYPE_MAP = {
    'share':         'Акции',
    'bond':          'Облигации',
    'etf':           'Фонды',
    'precious_metal':'Драг. металлы',
    'real_estate':   'Недвижимость',
    'currency':      'Деньги',
    'crypto':        'Крипта',
}


def _type_label(raw: str) -> str:
    for key, label in TYPE_MAP.items():
        if key in raw.lower():
            return label
    if 'фонд' in raw.lower() or 'reit' in raw.lower():
        return 'Недвижимость'
    return 'Другие'


def _dark_theme():
    plt.rcParams.update({
        'figure.facecolor': _DARK_BG, 'axes.facecolor':  _PANEL_BG,
        'axes.edgecolor':   _GRID,    'axes.labelcolor':  _DIM,
        'axes.titlecolor':  _TEXT,    'xtick.color':      _DIM,
        'ytick.color':      _DIM,     'grid.color':       _GRID,
        'grid.linewidth':   0.6,      'text.color':       _TEXT,
        'font.family':      'DejaVu Sans',
        'legend.facecolor': _PANEL_BG,'legend.edgecolor': _GRID,
        'legend.labelcolor':_TEXT,
    })


def plot_assets(positions: List[Dict],
                out_path: str = 'assets_distribution.png') -> None:
    """
    Three-panel dark financial dashboard:
      Panel A (top-left)  — Donut: asset-class allocation
      Panel B (top-right) — Horizontal bars: top-15 positions
      Panel C (bottom)    — Stacked bars: per-account breakdown
    """
    _dark_theme()

    # Aggregate by asset class (all positions including cash)
    class_totals: Dict[str, float] = {}
    for p in positions:
        lbl = _type_label(p['instrument_type'])
        class_totals[lbl] = class_totals.get(lbl, 0.0) + float(p['rub_value'])
    class_totals = dict(sorted(class_totals.items(), key=lambda x: x[1], reverse=True))
    total_all = sum(class_totals.values()) or 1.0

    # Top-15 positions (exclude pure cash)
    investable = [p for p in positions
                  if _type_label(p['instrument_type']) != 'Деньги']
    asset_vals:  Dict[str, float] = {}
    asset_class: Dict[str, str]   = {}
    for p in investable:
        t = p['ticker']
        asset_vals[t]  = asset_vals.get(t, 0.0)  + float(p['rub_value'])
        asset_class[t] = _type_label(p['instrument_type'])
    total_inv = sum(asset_vals.values()) or 1.0
    top15 = sorted(asset_vals.items(), key=lambda x: x[1], reverse=True)[:15]

    # Per-account breakdown
    accounts   = list(dict.fromkeys(p['account'] for p in positions))
    all_classes = list(class_totals.keys())
    acct_data: Dict[str, Dict[str, float]] = {a: {} for a in accounts}
    for p in positions:
        a = p['account']
        c = _type_label(p['instrument_type'])
        acct_data[a][c] = acct_data[a].get(c, 0.0) + float(p['rub_value'])

    # ── Figure layout ─────────────────────────────────────────────────────
    fig = plt.figure(figsize=(22, 14), facecolor=_DARK_BG)
    fig.text(0.5, 0.97, 'PORTFOLIO  DISTRIBUTION',
             ha='center', fontsize=20, fontweight='bold',
             color=_TEXT, family='monospace')
    fig.text(0.5, 0.935, f'Total NAV  ·  {total_all:,.0f} ₽',
             ha='center', fontsize=11, color=_DIM)

    gs = GridSpec(2, 3, figure=fig,
                  left=0.04, right=0.97, top=0.91, bottom=0.06,
                  wspace=0.38, hspace=0.50)

    ax_donut = fig.add_subplot(gs[0, 0])
    ax_bar   = fig.add_subplot(gs[0, 1:])
    ax_acct  = fig.add_subplot(gs[1, :])

    # ── A: Donut ──────────────────────────────────────────────────────────
    labels  = list(class_totals.keys())
    sizes   = [class_totals[l] for l in labels]
    colours = [TYPE_PALETTE.get(l, '#484F58') for l in labels]

    wedges, _, autotexts = ax_donut.pie(
        sizes, labels=None, colors=colours,
        autopct=lambda p: f'{p:.1f}%' if p > 3 else '',
        pctdistance=0.76,
        wedgeprops={'linewidth': 2.5, 'edgecolor': _DARK_BG, 'width': 0.50},
        startangle=90, counterclock=False,
    )
    for at in autotexts:
        at.set_fontsize(8.5); at.set_color(_TEXT); at.set_fontweight('bold')

    ax_donut.text(0,  0.10, 'NAV',         ha='center', color=_DIM,  fontsize=9)
    ax_donut.text(0, -0.15, f'{total_all/1e6:.1f}M ₽',
                  ha='center', color=_TEXT, fontsize=12, fontweight='bold')

    patches = [mpatches.Patch(color=TYPE_PALETTE.get(l, '#484F58'),
                               label=f'{l}  {v/total_all*100:.1f}%')
               for l, v in class_totals.items()]
    ax_donut.legend(handles=patches, loc='lower center',
                    bbox_to_anchor=(0.5, -0.32), ncol=2,
                    fontsize=8, framealpha=0, handlelength=0.8,
                    handletextpad=0.5, labelspacing=0.35)
    ax_donut.set_title('Asset Class Allocation',
                       color=_TEXT, fontsize=11, fontweight='bold', pad=12)

    # ── B: Top-15 bars ────────────────────────────────────────────────────
    if top15:
        tickers_r = [t for t, _ in reversed(top15)]
        values_r  = [v for _, v in reversed(top15)]
        bar_cols  = [TYPE_PALETTE.get(asset_class.get(t, 'Другие'), '#484F58')
                     for t in tickers_r]

        bars = ax_bar.barh(range(len(tickers_r)), values_r,
                           color=bar_cols, height=0.62, edgecolor='none')
        ax_bar.set_yticks(range(len(tickers_r)))
        ax_bar.set_yticklabels(tickers_r, fontsize=8.5, color=_TEXT)
        ax_bar.set_xlabel('Value (RUB)', color=_DIM, fontsize=9)
        ax_bar.set_title('Top-15 Positions',
                         color=_TEXT, fontsize=11, fontweight='bold', pad=12)
        ax_bar.xaxis.set_major_formatter(
            mticker.FuncFormatter(
                lambda x, _: f'{x/1e6:.1f}M' if x >= 1e6 else f'{x/1e3:.0f}K'))
        ax_bar.grid(axis='x', linestyle='--', alpha=0.25)
        ax_bar.spines[['top', 'right']].set_visible(False)

        max_v = max(values_r) if values_r else 1
        for bar, val in zip(bars, values_r):
            pct = val / total_inv * 100
            ax_bar.text(
                val + max_v * 0.01,
                bar.get_y() + bar.get_height() / 2,
                f'{val/1e3:.0f}K  {pct:.1f}%',
                va='center', ha='left', fontsize=7.5, color=_DIM,
            )

    # ── C: Per-account stacked bars ───────────────────────────────────────
    x       = np.arange(len(accounts))
    bottoms = np.zeros(len(accounts))

    for cls in all_classes:
        vals = np.array([acct_data[a].get(cls, 0.0) for a in accounts])
        ax_acct.bar(x, vals, 0.45, bottom=bottoms,
                    color=TYPE_PALETTE.get(cls, '#484F58'),
                    label=cls, edgecolor='none')
        for xi, (v, b) in enumerate(zip(vals, bottoms)):
            if v / total_all > 0.025:
                ax_acct.text(xi, b + v / 2,
                             f'{v/1e3:.0f}K',
                             ha='center', va='center',
                             fontsize=7, color='white', fontweight='bold')
        bottoms += vals

    ax_acct.set_xticks(x)
    ax_acct.set_xticklabels(accounts, fontsize=9, color=_TEXT)
    ax_acct.set_ylabel('Value (RUB)', color=_DIM, fontsize=9)
    ax_acct.set_title('Per-Account Breakdown',
                      color=_TEXT, fontsize=11, fontweight='bold', pad=12)
    ax_acct.yaxis.set_major_formatter(
        mticker.FuncFormatter(
            lambda x, _: f'{x/1e6:.1f}M' if x >= 1e6 else f'{x/1e3:.0f}K'))
    ax_acct.legend(loc='upper right', fontsize=8, framealpha=0.25,
                   ncol=min(len(all_classes), 6))
    ax_acct.grid(axis='y', linestyle='--', alpha=0.25)
    ax_acct.spines[['top', 'right']].set_visible(False)

    fig.text(0.97, 0.01, 'T-Invest Portfolio Analytics',
             ha='right', color=_DIM, fontsize=8, style='italic')

    fig.savefig(out_path, dpi=180, bbox_inches='tight', facecolor=_DARK_BG)
    plt.close(fig)
    logger.info(f"Графики сохранены: {out_path}")
    print(f"✓ Charts → {out_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# DCF TEMPLATE
# ═══════════════════════════════════════════════════════════════════════════════

def print_dcf_template():
    print("""
╔══════════════════════════════════════════════════════════════════════════════╗
║              DISCOUNTED CASH FLOW  —  ANALYSIS TEMPLATE                     ║
╚══════════════════════════════════════════════════════════════════════════════╝

─────────────────────────────────────────────────────────────────────────────
STEP 1  ·  COMPANY INPUTS
─────────────────────────────────────────────────────────────────────────────
  ticker         = "SIBN"            # TODO: change per company
  name           = "Газпромнефть"
  shares_out     = 4_741_299_639     # shares outstanding
  net_debt       = 620e9             # RUB: total debt - cash
  minority_int   = 0                 # from balance sheet

─────────────────────────────────────────────────────────────────────────────
STEP 2  ·  HISTORICAL FREE CASH FLOW  (RUB, last 5 years)
─────────────────────────────────────────────────────────────────────────────
  historical_fcf = {
      2020: 180e9,   # TODO: fill from annual report
      2021: 310e9,
      2022: 290e9,
      2023: 340e9,
      2024: 0,       # TODO: update
  }

─────────────────────────────────────────────────────────────────────────────
STEP 3  ·  PROJECTION ASSUMPTIONS
─────────────────────────────────────────────────────────────────────────────
  wacc             = 0.16    # OFZ yield + equity risk premium + beta adj.
  terminal_growth  = 0.04    # long-run nominal GDP growth (Russia ~4%)
  projection_years = 5
  growth_rates     = [0.08, 0.07, 0.05, 0.04, 0.04]  # y1…y5

─────────────────────────────────────────────────────────────────────────────
STEP 4  ·  DCF ENGINE
─────────────────────────────────────────────────────────────────────────────

import numpy as np

def run_dcf(base_fcf, growth_rates, wacc, terminal_growth,
            net_debt, shares_out):
    fcf        = base_fcf
    projected  = []
    for g in growth_rates:
        fcf *= (1 + g)
        projected.append(fcf)

    years            = len(projected)
    disc_factors     = [1 / (1 + wacc) ** t for t in range(1, years + 1)]
    pv_fcfs          = [cf * df for cf, df in zip(projected, disc_factors)]
    terminal_value   = projected[-1] * (1 + terminal_growth) / (wacc - terminal_growth)
    pv_terminal      = terminal_value * disc_factors[-1]

    enterprise_value = sum(pv_fcfs) + pv_terminal
    equity_value     = enterprise_value - net_debt
    intrinsic_share  = equity_value / shares_out
    return {
        "intrinsic_per_share": intrinsic_share,
        "enterprise_value":    enterprise_value,
        "pv_fcfs":             sum(pv_fcfs),
        "pv_terminal":         pv_terminal,
        "tv_share":            pv_terminal / enterprise_value,
    }

# base = max(historical_fcf.values())
# result = run_dcf(base, growth_rates, wacc, terminal_growth, net_debt, shares_out)
# upside = result["intrinsic_per_share"] / current_price - 1

─────────────────────────────────────────────────────────────────────────────
STEP 5  ·  SENSITIVITY TABLE  (WACC × terminal growth)
─────────────────────────────────────────────────────────────────────────────

import pandas as pd

# def sensitivity_table(base_fcf, growth_rates, net_debt, shares_out,
#                       wacc_range, tg_range):
#     rows = {}
#     for w in wacc_range:
#         row = {}
#         for g in tg_range:
#             r = run_dcf(base_fcf, growth_rates, w, g, net_debt, shares_out)
#             row[f"tg={g:.1%}"] = round(r["intrinsic_per_share"], 0)
#         rows[f"wacc={w:.1%}"] = row
#     return pd.DataFrame(rows).T
#
# wacc_range = np.arange(0.13, 0.21, 0.01)
# tg_range   = np.arange(0.02, 0.07, 0.01)
# print(sensitivity_table(base, growth_rates, net_debt, shares_out,
#                         wacc_range, tg_range))

─────────────────────────────────────────────────────────────────────────────
STEP 6  ·  PORTFOLIO INTEGRATION
─────────────────────────────────────────────────────────────────────────────

# After running run_dcf() and plot_correlation():
#   upside_pct  = (intrinsic_per_share / current_price - 1) * 100
#   weight      = position_rub_value / total_portfolio_rub
#   contribution= weight * upside_pct
#
# Correlation check: if the asset is highly correlated with existing
# holdings (r > 0.7), the marginal diversification benefit is low —
# require higher upside to justify the position size.

─────────────────────────────────────────────────────────────────────────────
TIPS
─────────────────────────────────────────────────────────────────────────────
  • FCF = Operating CF - CapEx (from cash flow statement)
  • OFZ 10y as risk-free rate  →  cbr.ru
  • Beta: regress stock daily returns on IMOEX returns (252d window)
  • For oil/gas: run scenario tree on Brent price, not a single FCF path
  • TV typically 60–80% of EV — if >85%, the model is very sensitive;
    stress-test terminal_growth aggressively
""")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_corr_arg(arg: str):
    """Parse correlation CLI modifiers: top15, min1, spearman."""
    if arg == 'spearman':
        return {'corr_method': 'spearman'}
    if arg.startswith('top') and arg[3:].isdigit():
        return {'top_n': int(arg[3:])}
    if arg.startswith('min') and arg[3:].replace('.', '', 1).isdigit():
        return {'min_weight_pct': float(arg[3:])}
    return {}


def main():
    out_path = 'Finance_updated.xlsx'
    command  = None
    days     = 180
    corr_kw: Dict = {}

    for arg in sys.argv[1:]:
        if arg.endswith('.xlsx'):
            out_path = arg
        elif arg in ('charts', 'correlation', 'dcf'):
            command = arg
        elif arg.isdigit():
            days = int(arg)
        elif command == 'correlation':
            corr_kw.update(_parse_corr_arg(arg))

    if command == 'dcf':
        print_dcf_template()
        return

    logger.info("Подключение к T-Invest API…")
    with Client(TOKEN) as client:
        rates    = get_currency_rates(client)
        api_rows = fetch_all_positions(client, rates)
    logger.info(f"Получено позиций: {len(api_rows)}")

    if command == 'charts':
        plot_assets(api_rows)
        return

    if command == 'correlation':
        run_correlation_analysis(
            api_rows,
            token=TOKEN,
            days=days,
            **corr_kw,
        )
        return

    # Default: update/create Excel, then render charts
    if os.path.exists(out_path):
        logger.info(f"Режим ОБНОВЛЕНИЯ: {out_path}")
        update_workbook(out_path, api_rows)
    else:
        logger.info(f"Режим СОЗДАНИЯ: {out_path}")
        wb = build_workbook(api_rows)
        wb.save(out_path)
        logger.info(f"Файл создан: {out_path}")

    plot_assets(api_rows)


if __name__ == '__main__':
    main()
